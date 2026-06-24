# ============================================================
#  ENCORI SCRAPER - FINAL VERSION
#  Reads p-value and HR directly from the KIRP survival graph
# ============================================================

import time
import os
import re
import pandas as pd
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

EXCEL_FILE = "Updated_Preliminary_Analysis.xlsx"
SHEETS = ["U FDR", "D FDR"]
GRAPHS_FOLDER = "graphs"
ENCORI_URL = "https://rnasysu.com/encori/panGeneSurvivalExp.php"

if not os.path.exists(GRAPHS_FOLDER):
    os.makedirs(GRAPHS_FOLDER)

print("\n--- Reading Excel file ---")
excel_data = {}
for sheet in SHEETS:
    df = pd.read_excel(EXCEL_FILE, sheet_name=sheet)
    df['Gene_ID_clean'] = df['Gene_ID'].astype(str).str.split('.').str[0]
    excel_data[sheet] = df
    print(f"  Sheet '{sheet}': {len(df)} genes found")

print("\n--- Setting up Chrome ---")
options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
options.add_argument("--disable-notifications")
driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install()),
    options=options
)
wait = WebDriverWait(driver, 20)

def scrape_gene(gene_id_clean, gene_id_original, sheet_name):
    print(f"\n  [{sheet_name}] Searching: {gene_id_clean}")

    try:
        driver.get(ENCORI_URL)
        time.sleep(2)

        # Type gene ID in search box
        gene_input = wait.until(
            EC.presence_of_element_located((By.XPATH,
                "//input[@placeholder='PTEN']"
            ))
        )
        gene_input.clear()
        gene_input.send_keys(gene_id_clean)
        time.sleep(0.5)

        # Click Search
        search_btn = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(text(),'Search')]"))
        )
        search_btn.click()
        time.sleep(4)

        # Select Kidney Renal Papillary Cell Carcinoma in the Cancer dropdown
        try:
            cancer_select = wait.until(
                EC.presence_of_element_located((By.XPATH, "//select[contains(@class,'cancer') or contains(@id,'cancer') or contains(@name,'cancer')]"))
            )
            for option in Select(cancer_select).options:
                if "Papillary" in option.text:
                    Select(cancer_select).select_by_visible_text(option.text)
                    print(f"    Selected: {option.text}")
                    break
        except:
            # Try any select dropdown that has Papillary option
            try:
                all_selects = driver.find_elements(By.TAG_NAME, "select")
                for sel in all_selects:
                    opts = Select(sel).options
                    for opt in opts:
                        if "Papillary" in opt.text:
                            Select(sel).select_by_visible_text(opt.text)
                            print(f"    Selected: {opt.text}")
                            break
            except Exception as e:
                print(f"    Could not select cancer: {e}")

        time.sleep(4)  # Wait for graph to update

        # Read p-value and HR from the graph text
        p_value = None
        hr_value = None

        try:
            # The graph shows text like:
            # "Log-Rank p=0.22"
            # "Hazard Ratio=1.45"
            page_source = driver.page_source

            # Extract Log-Rank p value
            p_match = re.search(r'Log-Rank p[=\s]*([0-9.eE+\-]+)', page_source)
            if p_match:
                p_value = float(p_match.group(1))
                print(f"    p-value: {p_value}")

            # Extract Hazard Ratio
            hr_match = re.search(r'Hazard Ratio[=\s]*([0-9.eE+\-]+)', page_source)
            if hr_match:
                hr_value = float(hr_match.group(1))
                print(f"    HR: {hr_value}")

            if p_value is None and hr_value is None:
                print(f"    No KIRP data found for {gene_id_clean}")

        except Exception as e:
            print(f"    Could not extract values: {e}")

        # Save screenshot
        safe_name = gene_id_original.replace('.', '_').replace('/', '_')
        screenshot_path = os.path.join(GRAPHS_FOLDER, f"{sheet_name}_{safe_name}.png")
        driver.save_screenshot(screenshot_path)
        print(f"    Screenshot saved")

        return p_value, hr_value

    except Exception as e:
        print(f"    FAILED: {e}")
        return None, None

# ============================================================
# MAIN LOOP
# ============================================================

print("\n--- Starting gene search ---")
results = {sheet: {} for sheet in SHEETS}

for sheet in SHEETS:
    df = excel_data[sheet]
    total = len(df)
    print(f"\n{'='*50}")
    print(f"Processing: {sheet} ({total} genes)")
    print(f"{'='*50}")

    for idx, (i, row) in enumerate(df.iterrows()):
        gene_original = str(row['Gene_ID'])
        gene_clean = str(row['Gene_ID_clean'])
        print(f"\n[{idx+1}/{total}] {gene_original}")

        p_val, hr_val = scrape_gene(gene_clean, gene_original, sheet)
        results[sheet][gene_original] = (p_val, hr_val)
        time.sleep(1)

driver.quit()
print("\n--- Browser closed ---")

# ============================================================
# WRITE RESULTS TO EXCEL
# ============================================================

print("\n--- Writing results to Excel ---")
wb = load_workbook(EXCEL_FILE)

for sheet in SHEETS:
    ws = wb[sheet]
    p_col = None
    hr_col = None

    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val == 'p-value':
            p_col = col
        elif val == 'HR':
            hr_col = col

    print(f"\nSheet '{sheet}': p-value=col{p_col}, HR=col{hr_col}")

    filled = 0
    for row in range(2, ws.max_row + 1):
        gene_id = str(ws.cell(row=row, column=1).value)
        if gene_id in results[sheet]:
            p_val, hr_val = results[sheet][gene_id]
            if p_val is not None and p_col:
                ws.cell(row=row, column=p_col).value = p_val
            if hr_val is not None and hr_col:
                ws.cell(row=row, column=hr_col).value = hr_val
            if p_val is not None or hr_val is not None:
                filled += 1

    print(f"  Filled {filled} genes")

output_file = "Updated_Preliminary_Analysis_WITH_ENCORI.xlsx"
wb.save(output_file)

print(f"\n{'='*50}")
print(f"Excel saved: {output_file}")
print(f"Graphs in: {GRAPHS_FOLDER}/")
print(f"{'='*50}")

for sheet in SHEETS:
    total = len(results[sheet])
    found_p  = sum(1 for v in results[sheet].values() if v[0] is not None)
    found_hr = sum(1 for v in results[sheet].values() if v[1] is not None)
    not_found = [g for g, v in results[sheet].items() if v[0] is None and v[1] is None]
    print(f"\n{sheet}: {total} genes | p-values: {found_p} | HR: {found_hr} | no data: {len(not_found)}")

print("\n=== ALL DONE ===")
